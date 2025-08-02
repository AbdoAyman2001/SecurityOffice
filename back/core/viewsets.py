from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import io
from .models import (
    PeopleHistory, CompaniesHistory, EmploymentHistory, FamilyRelationships,
    CorrespondenceTypes, Contacts, Correspondence,
    Attachments, CorrespondenceStatusLog, Permits, ApprovalDecisions,
    Accidents, Relocation, RelocationPeriod, Vehicle, CarPermit,
    CardPermits, CardPhotos
)
from .serializers.auth_serializers import UserSerializer
from .serializers.people_serializers import (
    PeopleHistorySerializer, CompaniesHistorySerializer,
    EmploymentHistorySerializer, FamilyRelationshipsSerializer,
    PeopleHistorySummarySerializer
)
from .serializers.correspondence_serializers import (
    CorrespondenceTypesSerializer, ContactsSerializer, CorrespondenceSerializer,
    AttachmentsSerializer, CorrespondenceSummarySerializer
)
from .serializers.permits_serializers import (
    PermitsSerializer, ApprovalDecisionsSerializer
)
from .serializers.accidents_serializers import AccidentsSerializer
from .serializers.relocation_serializers import (
    RelocationSerializer, RelocationPeriodSerializer
)
from .serializers.vehicles_serializers import (
    VehicleSerializer, CarPermitSerializer
)
from .serializers.cards_serializers import (
    CardPermitsSerializer, CardPhotosSerializer
)

User = get_user_model()


# ====================================== USER VIEWSETS ======================================
class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['username', 'email', 'first_name', 'last_name']
    filterset_fields = ['is_active', 'is_staff']
    ordering_fields = ['username', 'date_joined']
    ordering = ['username']


# ====================================== PEOPLE VIEWSETS ======================================
class PeopleHistoryViewSet(viewsets.ModelViewSet):
    queryset = PeopleHistory.objects.all()
    serializer_class = PeopleHistorySerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['full_name_arabic', 'full_name_english', 'national_id', 'person_guid']
    filterset_fields = ['is_current', 'nationality', 'alive', 'version']
    ordering_fields = ['full_name_arabic', 'start_date', 'version']
    ordering = ['-start_date']

    @action(detail=False, methods=['get'])
    def current_only(self, request):
        """Get only current versions of people records"""
        current_people = self.queryset.filter(is_current=True)
        serializer = PeopleHistorySummarySerializer(current_people, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        """Get all versions of a person's history by person_guid"""
        person = self.get_object()
        history = PeopleHistory.objects.filter(person_guid=person.person_guid).order_by('-version')
        serializer = self.get_serializer(history, many=True)
        return Response(serializer.data)


class CompaniesHistoryViewSet(viewsets.ModelViewSet):
    queryset = CompaniesHistory.objects.all()
    serializer_class = CompaniesHistorySerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['company_name', 'company_type']
    filterset_fields = ['is_current', 'company_type']
    ordering_fields = ['company_name', 'start_date']
    ordering = ['company_name']

    @action(detail=False, methods=['get'])
    def current_only(self, request):
        """Get only current versions of company records"""
        current_companies = self.queryset.filter(is_current=True)
        serializer = self.get_serializer(current_companies, many=True)
        return Response(serializer.data)


class EmploymentHistoryViewSet(viewsets.ModelViewSet):
    queryset = EmploymentHistory.objects.all()
    serializer_class = EmploymentHistorySerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['person_guid', 'job_title', 'company__company_name']
    filterset_fields = ['still_hired', 'is_current', 'company']
    ordering_fields = ['start_date', 'job_title']
    ordering = ['-start_date']


class FamilyRelationshipsViewSet(viewsets.ModelViewSet):
    queryset = FamilyRelationships.objects.all()
    serializer_class = FamilyRelationshipsSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['worker_person_guid', 'family_member_person_guid']
    filterset_fields = ['relationship_type', 'status']
    ordering_fields = ['relationship_type']
    ordering = ['relationship_type']


# ====================================== CORRESPONDENCE VIEWSETS ======================================
class CorrespondenceTypesViewSet(viewsets.ModelViewSet):
    queryset = CorrespondenceTypes.objects.all()
    serializer_class = CorrespondenceTypesSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['type_name']
    ordering = ['type_name']


class ContactsViewSet(viewsets.ModelViewSet):
    queryset = Contacts.objects.all()
    serializer_class = ContactsSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name']
    filterset_fields = ['contact_type', 'is_approver']
    ordering_fields = ['name', 'contact_type']
    ordering = ['name']

    @action(detail=False, methods=['get'])
    def approvers(self, request):
        """Get only contacts that are approvers"""
        approvers = self.queryset.filter(is_approver=True)
        serializer = self.get_serializer(approvers, many=True)
        return Response(serializer.data)


class CorrespondenceViewSet(viewsets.ModelViewSet):
    queryset = Correspondence.objects.select_related(
        'type', 'contact', 'current_status', 'assigned_to', 'parent_correspondence'
    ).all()
    serializer_class = CorrespondenceSerializer
    authentication_classes = [TokenAuthentication]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    # Enhanced search fields for better search functionality
    search_fields = [
        'reference_number', 'subject', 'summary',
        'contact__name', 'type__type_name', 'assigned_to__full_name_arabic',
        'assigned_to__username', 'current_status__procedure_name'
    ]
    
    # Fixed filterset_fields to support frontend filtering with __in lookups
    filterset_fields = {
        # Direct field filters
        'direction': ['exact', 'in'],
        'priority': ['exact', 'in'], 
        'correspondence_date': ['exact', 'gte', 'lte', 'in'],
        'reference_number': ['exact', 'icontains', 'in'],
        'created_at': ['exact', 'gte', 'lte'],
        'updated_at': ['exact', 'gte', 'lte'],
        
        # Related field filters with __in support for frontend
        'type__type_name': ['exact', 'in', 'icontains'],
        'contact__name': ['exact', 'in', 'icontains'],
        'current_status__procedure_name': ['exact', 'in', 'icontains'],
        'assigned_to__full_name_arabic': ['exact', 'in', 'icontains'],
        'assigned_to__username': ['exact', 'in', 'icontains'],
        'parent_correspondence__reference_number': ['exact', 'in', 'icontains'],
    }
    
    # Enhanced ordering fields to support all columns
    ordering_fields = [
        'correspondence_id', 'correspondence_date', 'reference_number', 
        'priority', 'direction', 'subject', 'created_at', 'updated_at',
        'type__type_name', 'contact__name', 'current_status__procedure_name',
        'assigned_to__full_name_arabic', 'parent_correspondence__reference_number'
    ]
    ordering = ['-correspondence_date']
    
    def get_permissions(self):
        """Set permissions based on action"""
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            # Require authentication for write operations
            permission_classes = [IsAuthenticated]
        else:
            # Allow read operations without authentication
            permission_classes = []
        return [permission() for permission in permission_classes]
    
    def create(self, request, *args, **kwargs):
        """Create correspondence with related records"""
        from django.db import transaction
        from django.contrib.auth.models import AnonymousUser
        
        # Check if user is authenticated
        if isinstance(request.user, AnonymousUser):
            return Response(
                {'error': 'Authentication required to create correspondence'},
                status=status.HTTP_401_UNAUTHORIZED
            )
        
        # Extract related data from request
        contact_id = request.data.get('contact')
        attachments_data = request.data.get('attachments', [])
        
        with transaction.atomic():
            # Create the main correspondence record
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            correspondence = serializer.save()
            
            # Contact is now handled directly in the correspondence model
            # No need to create separate correspondence_contact records
            
            # Create initial status log entry if user is authenticated and status exists
            if correspondence.current_status and request.user.is_authenticated:
                try:
                    CorrespondenceStatusLog.objects.create(
                        correspondence=correspondence,
                        form_status_name=None,  # Initial status
                        to_status_name=correspondence.current_status.procedure_name if correspondence.current_status else None,
                        changed_by=request.user,
                        change_reason='Initial correspondence creation'
                    )
                except Exception as e:
                    # Continue without failing the entire operation
                    pass
            
            # Return the created correspondence with all related data
            headers = self.get_success_headers(serializer.data)
            return Response(
                {
                    'correspondence': serializer.data,
                    'message': 'Correspondence created successfully'
                },
                status=status.HTTP_201_CREATED,
                headers=headers
            )

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get correspondence summary for listings"""
        queryset = self.filter_queryset(self.get_queryset())
        
        # Apply pagination
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = CorrespondenceSummarySerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = CorrespondenceSummarySerializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export all correspondence data to Excel without pagination"""
        # Get all filtered data without pagination, sorted by ID ascending
        queryset = self.filter_queryset(self.get_queryset()).order_by('correspondence_id')
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Russian Letters"
        
        # Define headers based on the frontend column configuration
        headers = [
            'كود الخطاب',  # correspondence_id
            'الرقم المرجعي',   # reference_number
            'تاريخ الخطاب',  # correspondence_date
            'جهة المخاطبة',    # contact
            'النوع',         # type
            'الموضوع',       # subject
            'الاتجاه',       # direction
            'الأولوية',      # priority
            'الحالة الحالية', # current_status
            'مُكلف إلى',     # assigned_to
            'ملخص الخطاب',        # summary
            'المرفقات',      # attachments
            'تاريخ الإضافة إلى المنظومة',  # created_at
            'تاريخ آخر تحديث'   # updated_at
        ]
        
        # Style the header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        for row_num, correspondence in enumerate(queryset, 2):
            # Format data according to the frontend display logic
            row_data = [
                correspondence.correspondence_id,
                correspondence.reference_number or 'غير محدد',
                correspondence.correspondence_date.strftime('%d/%m/%Y') if correspondence.correspondence_date else 'غير محدد',
                correspondence.contact.name if correspondence.contact else 'غير محدد',
                correspondence.type.type_name if correspondence.type else 'غير محدد',
                correspondence.subject or 'غير محدد',
                self._get_direction_label(correspondence.direction),
                self._get_priority_label(correspondence.priority),
                correspondence.current_status.procedure_name if correspondence.current_status else 'غير محدد',
                correspondence.assigned_to.full_name_arabic if correspondence.assigned_to and hasattr(correspondence.assigned_to, 'full_name_arabic') else (correspondence.assigned_to.username if correspondence.assigned_to else 'غير مُكلف'),
                correspondence.summary or 'لا يوجد ملخص',
                self._get_attachment_names(correspondence),
                correspondence.created_at.strftime('%d/%m/%Y %H:%M') if correspondence.created_at else 'غير محدد',
                correspondence.updated_at.strftime('%d/%m/%Y %H:%M') if correspondence.updated_at else 'غير محدد'
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response with Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Generate filename with current date
        filename = f"russian_letters_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save workbook to response
        wb.save(response)
        return response
    
    def _get_direction_label(self, direction):
        """Convert direction value to Arabic label"""
        direction_map = {
            'Incoming': 'وارد',
            'Outgoing': 'صادر',
            'Internal': 'داخلي'
        }
        return direction_map.get(direction, direction or 'غير محدد')
    
    def _get_priority_label(self, priority):
        """Convert priority value to Arabic label"""
        priority_map = {
            'high': 'عالية',
            'normal': 'عادية',
            'low': 'منخفضة'
        }
        return priority_map.get(priority, priority or 'عادية')
    
    def _get_attachment_names(self, correspondence):
        """Get attachment names as comma-separated list"""
        try:
            if hasattr(correspondence, 'attachments'):
                attachment_names = []
                for attachment in correspondence.attachments.all():
                    # Use file_name if available, otherwise use the filename from the file field
                    name = getattr(attachment, 'file_name', None) or getattr(attachment, 'original_filename', None)
                    if name:
                        attachment_names.append(name)
                    elif hasattr(attachment, 'file') and attachment.file:
                        # Extract filename from file path
                        import os
                        attachment_names.append(os.path.basename(attachment.file.name))
                
                if attachment_names:
                    return '\n'.join(attachment_names)
                else:
                    return 'لا توجد مرفقات'
            else:
                return 'لا توجد مرفقات'
        except Exception as e:
            # Fallback in case of any error
            return 'خطأ في تحميل المرفقات'
    
    @action(detail=True, methods=['get'], url_path='detail-with-relations')
    def detail_with_relations(self, request, pk=None):
        """Get comprehensive letter detail with all related data in one call"""
        try:
            # Get the main correspondence with all related data
            correspondence = self.get_object()
            
            # Serialize the main correspondence
            serializer = self.get_serializer(correspondence)
            letter_data = serializer.data
            
            # Get status history
            status_logs = CorrespondenceStatusLog.objects.filter(
                correspondence=correspondence
            ).order_by('-created_at')
            from .serializers.correspondence_serializers import CorrespondenceStatusLogSerializer
            status_history = CorrespondenceStatusLogSerializer(status_logs, many=True).data
            
            # Get related correspondence (children, siblings, parent)
            children = Correspondence.objects.filter(
                parent_correspondence=correspondence
            ).select_related('type', 'contact', 'current_status')
            
            siblings = []
            if correspondence.parent_correspondence:
                siblings = Correspondence.objects.filter(
                    parent_correspondence=correspondence.parent_correspondence
                ).exclude(
                    correspondence_id=correspondence.correspondence_id
                ).select_related('type', 'contact', 'current_status')
            
            parent = []
            if correspondence.parent_correspondence:
                parent = [correspondence.parent_correspondence]
            
            # Combine all related correspondence
            all_related = list(children) + list(siblings) + parent
            related_correspondence = CorrespondenceSerializer(all_related, many=True).data if all_related else []
            
            # Get correspondence types
            correspondence_types = CorrespondenceTypes.objects.all()
            types_data = CorrespondenceTypesSerializer(correspondence_types, many=True).data
            
            # Get contacts
            contacts = Contacts.objects.all()
            from .serializers.correspondence_serializers import ContactsSerializer
            contacts_data = ContactsSerializer(contacts, many=True).data
            
            # Get procedures for this correspondence type
            procedures_data = []
            if correspondence.type:
                from .models import CorrespondenceTypeProcedure
                from .serializers.correspondence_serializers import CorrespondenceTypeProcedureSerializer
                procedures = CorrespondenceTypeProcedure.objects.filter(
                    correspondence_type=correspondence.type
                ).order_by('procedure_order')
                procedures_data = CorrespondenceTypeProcedureSerializer(procedures, many=True).data
            
            # Prepare response data
            response_data = {
                'letter': letter_data,
                'status_history': status_history,
                'related_correspondence': related_correspondence,
                'correspondence_types': types_data,
                'contacts': contacts_data,
                'procedures': procedures_data,
                'metadata': {
                    'has_parent': correspondence.parent_correspondence is not None,
                    'children_count': len(list(children)),
                    'siblings_count': len(list(siblings)),
                    'total_related': len(related_correspondence)
                }
            }
            
            return Response(response_data, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response(
                {'error': f'Failed to fetch letter details: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class AttachmentsViewSet(viewsets.ModelViewSet):
    queryset = Attachments.objects.all()
    serializer_class = AttachmentsSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['file_name']
    filterset_fields = ['file_type', 'correspondence']
    ordering_fields = ['file_name', 'file_size']
    ordering = ['file_name']
    
    @action(detail=False, methods=['post'])
    def upload(self, request):
        """Upload files for a correspondence"""
        import os
        from django.conf import settings
        from django.core.files.storage import default_storage
        from django.core.files.base import ContentFile
        import mimetypes
        correspondence_id = request.data.get('correspondence_id')
        files = request.FILES.getlist('files')
        
        if not correspondence_id:
            return Response(
                {'error': 'correspondence_id is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not files:
            return Response(
                {'error': 'No files provided'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            correspondence = Correspondence.objects.get(correspondence_id=correspondence_id)
        except Correspondence.DoesNotExist:
            return Response(
                {'error': 'Correspondence not found'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        uploaded_files = []
        
        for file in files:
            # Create attachment record with FileField handling the file storage
            attachment = Attachments.objects.create(
                correspondence=correspondence,
                file=file,  # FileField handles the upload path and storage
                file_name=file.name,
                file_type=file.content_type,
                file_size=file.size
            )
            
            uploaded_files.append({
                'attachment_id': attachment.attachment_id,
                'file_name': attachment.file_name,
                'file_size': attachment.file_size,
                'file_type': attachment.file_type,
                'file_url': attachment.file.url if attachment.file else None
            })
        
        return Response(
            {
                'message': f'Successfully uploaded {len(uploaded_files)} files',
                'files': uploaded_files
            },
            status=status.HTTP_201_CREATED
        )

    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        """Download an attachment file"""
        from django.http import FileResponse, Http404
        
        attachment = self.get_object()
        
        # Check if file exists using FileField
        if not attachment.file or not attachment.file.name:
            raise Http404("File not found")
        
        try:
            response = FileResponse(
                attachment.file.open('rb'),
                content_type=attachment.file_type or 'application/octet-stream',
                as_attachment=True,
                filename=attachment.file_name
            )
            return response
        except FileNotFoundError:
            raise Http404("File not found")


# ====================================== APPROVAL VIEWSETS ======================================
class PermitsViewSet(viewsets.ModelViewSet):
    queryset = Permits.objects.all()
    serializer_class = PermitsSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['person_guid', 'company__company_name']
    filterset_fields = ['permit_holder_type', 'permit_status', 'effective_date', 'expiry_date']
    ordering_fields = ['effective_date', 'expiry_date', 'permit_status']
    ordering = ['-effective_date']

    @action(detail=False, methods=['get'])
    def active(self, request):
        """Get only active permits"""
        active_permits = self.queryset.filter(permit_status='Active')
        serializer = self.get_serializer(active_permits, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def expiring_soon(self, request):
        """Get permits expiring in the next 30 days"""
        from datetime import date, timedelta
        expiry_threshold = date.today() + timedelta(days=30)
        expiring_permits = self.queryset.filter(
            permit_status='Active',
            expiry_date__lte=expiry_threshold,
            expiry_date__gte=date.today()
        )
        serializer = self.get_serializer(expiring_permits, many=True)
        return Response(serializer.data)


class ApprovalDecisionsViewSet(viewsets.ModelViewSet):
    queryset = ApprovalDecisions.objects.all()
    serializer_class = ApprovalDecisionsSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['decision_status', 'approver_contact', 'permit', 'decision_date']
    ordering_fields = ['decision_date', 'decision_status']
    ordering = ['-decision_date']


# ====================================== ACCIDENTS VIEWSETS ======================================
class AccidentsViewSet(viewsets.ModelViewSet):
    queryset = Accidents.objects.all()
    serializer_class = AccidentsSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['person_guid', 'description', 'address']
    filterset_fields = ['date']
    ordering_fields = ['date']
    ordering = ['-date']


# ====================================== RELOCATION VIEWSETS ======================================
class RelocationViewSet(viewsets.ModelViewSet):
    queryset = Relocation.objects.all()
    serializer_class = RelocationSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['person_guid']
    filterset_fields = ['approval_status', 'building_letter', 'building_number']
    ordering_fields = ['building_number', 'flat_number']
    ordering = ['building_number', 'building_letter', 'flat_number']


class RelocationPeriodViewSet(viewsets.ModelViewSet):
    queryset = RelocationPeriod.objects.all()
    serializer_class = RelocationPeriodSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['relocation', 'start_date', 'end_date']
    ordering_fields = ['start_date', 'end_date']
    ordering = ['-start_date']


# ====================================== VEHICLES VIEWSETS ======================================
class VehicleViewSet(viewsets.ModelViewSet):
    queryset = Vehicle.objects.all()
    serializer_class = VehicleSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['plate_number', 'vehicle_id']
    filterset_fields = ['organization', 'company', 'start_date', 'end_date']
    ordering_fields = ['vehicle_id', 'start_date']
    ordering = ['vehicle_id']


class CarPermitViewSet(viewsets.ModelViewSet):
    queryset = CarPermit.objects.all()
    serializer_class = CarPermitSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['vehicle', 'start_date', 'end_date']
    ordering_fields = ['start_date', 'end_date']
    ordering = ['-start_date']


# ====================================== CARD PERMITS VIEWSETS ======================================
class CardPermitsViewSet(viewsets.ModelViewSet):
    queryset = CardPermits.objects.all()
    serializer_class = CardPermitsSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['permit_number', 'person_guid']
    filterset_fields = ['permit_type', 'status', 'issue_date', 'expiration_date']
    ordering_fields = ['issue_date', 'expiration_date', 'permit_number']
    ordering = ['-issue_date']

    @action(detail=False, methods=['get'])
    def active(self, request):
        """Get only active card permits"""
        active_cards = self.queryset.filter(status='Active')
        serializer = self.get_serializer(active_cards, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def expiring_soon(self, request):
        """Get card permits expiring in the next 30 days"""
        from datetime import date, timedelta
        expiry_threshold = date.today() + timedelta(days=30)
        expiring_cards = self.queryset.filter(
            status='Active',
            expiration_date__lte=expiry_threshold,
            expiration_date__gte=date.today()
        )
        serializer = self.get_serializer(expiring_cards, many=True)
        return Response(serializer.data)


class CardPhotosViewSet(viewsets.ModelViewSet):
    queryset = CardPhotos.objects.all()
    serializer_class = CardPhotosSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['file_name']
    filterset_fields = ['permit', 'mime_type']
    ordering_fields = ['uploaded_at', 'file_name']
    ordering = ['-uploaded_at']
